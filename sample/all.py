import face_recognition
import os
import pickle
import serial
import ast
import numpy as np
import drivers
import time
import RPi.GPIO as gpio
gpio.setmode(gpio.BOARD)
buzzerpin = 40
gpio.setup(buzzerpin,gpio.OUT)
display = drivers.Lcd()
devices = ['/dev/ttyACM0','/dev/ttyACM1']
for device in devices:
 try:
  ser = serial.Serial(device, 9600 , timeout=1)
  ser.flush()
  break;
 except Exception as e :
     continue;
def collect_send_data(send_string) :
    print(send_string)
    ser.write(send_string.encode('utf-8'))
def find_ip():
    ip=os.popen('hostname -I').read()
    ip = ip.replace("\n","")
    if ip=="" :
     ip = ' NO INTERNET' 
    return ip
face_image = []
face_encodings = {}
phone_numbers = {}
attendees = []
with open('/home/pi/faceR_prj/encodings.dat','rb') as f:
    face_encodings = pickle.load(f)
known_faces = np.array(list(face_encodings.values()))     
known_face_names = list(face_encodings.keys())
time.sleep(5)
while(1):
 display.lcd_clear()
 display.lcd_display_string("SYSTEM ACTIVE", 1)
 display.lcd_display_string("YOUR IP:"+find_ip(),3)
 collect_send_data(',')
 while(1):
    if find_ip()==' NO INTERNET' :
      find_ip()
    message_recieved = ser.readline().decode('utf-8')
    if '1' in message_recieved:
        print(message_recieved)
        break
    if 's' in message_recieved:
        print(message_recieved)
        continue   
 display.lcd_clear()
 display.lcd_display_string("     SAY CHEESE!!", 2)
 os.system('raspistill -w 640 -h 480 -t 2000 -o /home/pi/faceR_prj/pic.jpg')
 unknown_image = face_recognition.load_image_file("/home/pi/faceR_prj/pic.jpg")
 display.lcd_clear()
 os.system('sudo rm -r /home/pi/faceR_prj/pic.jpg')
 try:
    unknown_face_encoding = face_recognition.face_encodings(unknown_image)[0]
 except IndexError:
    print("NO FACE DETECTED")
    display.lcd_clear()
    display.lcd_display_string("NO FACE DETECTED", 1)
    display.lcd_display_string("PLEASE TRY AGAIN", 2)
    collect_send_data(',')
    time.sleep(3)
    display.lcd_clear()
    continue
 matches = face_recognition.compare_faces(known_faces,unknown_face_encoding,0.46)
 print(matches)
 name = "Unknown"
 face_distances = face_recognition.face_distance(known_faces,unknown_face_encoding)
 best_match_index = np.argmin(face_distances)
 if matches[best_match_index]:
                name = known_face_names[best_match_index]           
 if name == "Unknown":
   display.lcd_clear()
   display.lcd_display_string("   UNREGISTED USER", 1)
   display.lcd_display_string("   PLEASE CONTACT " ,2)
   display.lcd_display_string("  YOUR LOCAL ADMIN", 3)
   collect_send_data(',')
   time.sleep(3)
 elif name in attendees:
     display.lcd_clear()
     display.lcd_display_string("   USER DATA  ", 1)
     display.lcd_display_string("  WAS PREVIOUSLY  " ,2)
     display.lcd_display_string("    LOGGED  " ,3)
     collect_send_data(',')
     time.sleep(3)
 else:  
   attendees.append(name)  
   display.lcd_clear()
   display.lcd_display_string("WELCOME :", 1)
   display.lcd_display_string(name.upper(), 2)
   time.sleep(2)
   display.lcd_clear()
   display.lcd_display_string("BRING YOUR ARM TO ", 1)
   display.lcd_display_string("MEASURE TEMPERATURE", 2)
   time.sleep(2)
   collect_send_data('*')
   flag=0
   for num in range(0,6):
     message_recieved = ser.readline().decode('utf-8')
     if message_recieved:
       flag=1 
       display.lcd_clear()
       LEN = len(message_recieved)
       message_recieved = message_recieved.replace(message_recieved[LEN-1],"")
       message_recieved = message_recieved.replace(message_recieved[LEN-2],"")       
       temperature = ast.literal_eval(message_recieved)
       attendees.append(name)
       NewName = name
       NewName = NewName.replace(" ","_")
       os.system('python3 /home/pi/faceR_prj/attendence.py '+NewName+' '+message_recieved)
       display.lcd_display_string("TEMPERATURE : "+message_recieved, 1)
       time.sleep(2)
       if temperature < 36 :
            sanflag=0
            display.lcd_clear()
            display.lcd_display_string("PLEASE SANITIZE", 1)
            display.lcd_display_string("BEFORE YOU LEAVE", 2)
            time.sleep(2)
            collect_send_data('#')
            for num in range(0,6):
              message_recieved = ser.readline().decode('utf-8')
              if message_recieved :
                 sanflag = 1
                 display.lcd_clear()
                 display.lcd_display_string("   HAVE A NICE",2)
                 display.lcd_display_string("       DAY",3)
                 time.sleep(2)
                 break
              else:   
                 numstr =str(num)
                 display.lcd_display_string("["+numstr+"/5s TILL TIME OUT]", 3)
                 time.sleep(1)
            if sanflag == 0:
                display.lcd_clear()
                display.lcd_display_string(" TIME OUT ",2)
                time.sleep(2)      
       else:
         display.lcd_clear()
         display.lcd_display_string("****WARNING****",1)
         display.lcd_display_string("TEMPERATURE VALUE",2)
         display.lcd_display_string("  EXCEEDS NORMS",3)
         display.lcd_display_string("  ACCESS DENIED",4)
         gpio.output(buzzerpin,gpio.HIGH)  
         collect_send_data('a')   
         time.sleep(5)
         gpio.output(buzzerpin,gpio.LOW)
         time.sleep(1)
       break
     else :
       numstr =str(num) 
       display.lcd_display_string("["+numstr+"/5s TILL TIME OUT]", 3)
       time.sleep(1)   
   if flag == 0:   
        display.lcd_clear()
        display.lcd_display_string("   TIME OUT", 2)
        time.sleep(2)
        collect_send_data('a')
        time.sleep(2)
 print(name)
Mailer.py
import email, smtplib, ssl
from datetime import date ,datetime
import pickle
import os
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
wb = openpyxl.Workbook()
TriggerTime = "12:14:00"
def Trigger():
    current_time = str(datetime.now().strftime("%H:%M:%S"))
    #print(current_time)
    if TriggerTime == current_time:
        return 1
    return 0
while(1):
 if Trigger() == 0 :
     continue      
 subject = "Today's attendence"
 body = "Data collected from covisafe"
 sender_email = ""    
 receiver_email = ""
 password = ""
 message = MIMEMultipart()
 message["From"] = sender_email
 message["To"] = receiver_email
 message["Subject"] = subject
 message["Bcc"] = receiver_email
 message.attach(MIMEText(body, "plain")
 sheet = wb.active
 temperatures={}
 phone_numbers={}
 try:
     with open('/home/pi/faceR_prj/temperature.dat','rb') as f:
         temperatures=pickle.load(f)
     with open('/home/pi/faceR_prj/phonenos.dat','rb') as P:
         phone_numbers=pickle.load(P)
     names = list(temperatures.keys())        
     FONT = Font(bold=True)
     sheet.cell(row=1, column=1).value="Name"
     sheet.cell(row=1, column=2).value="Phone Number"
     sheet.cell(row=1, column=3).value="Temperature"
     sheet['A1'].font = FONT
     sheet['B1'].font = FONT
     sheet['C1'].font = FONT
     i=2
     for name in names:
        sheet.cell(row=i, column=1).value=name
        sheet.cell(row=i, column=2).value=phone_numbers[name]
        sheet.cell(row=i, column=3).value=temperatures[name]
        i=i+1
     sheet.column_dimensions['A'].width = 20
     sheet.column_dimensions['B'].width = 20
     sheet.column_dimensions['C'].width = 20
 except :
      FONT = Font(bold=True)
      sheet.cell(row=1, column=1).value="N/A"
      sheet.cell(row=1, column=2).value="N/A"
      sheet.cell(row=1, column=3).value="N/A"
      sheet['A1'].font = FONT
      sheet['B1'].font = FONT
      sheet['C1'].font = FONT
 Date = str(date.today())
 wb.save("/home/pi/faceR_prj/Register/"+Date+".xlsx")
 filename = Date+".xlsx" 
 with open("/home/pi/faceR_prj/Register/"+filename, "rb") as attachment:
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())
 encoders.encode_base64(part)
 part.add_header(
    "Content-Disposition",
    f"attachment; filename= {filename}",
 )
 message.attach(part)
 text = message.as_string()
 context = ssl.create_default_context()
 try:
     with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
         server.login(sender_email, password)
         server.sendmail(sender_email, receiver_email, text)
         os.system("echo "" > /home/pi/faceR_prj/temperature.dat")
 except:
     print("No internet access. Unable to send mail")
Attendance.py
import sys
import pickle
arg1 = sys.argv[1].replace("_"," ")
arg2 = sys.argv[2]
attendee = {}
try:
            with open('/home/pi/faceR_prj/temperature.dat','rb') as f:
                        attendee=pickle.load(f)
                        attendee[arg1]=arg2                       
except:
            attendee[arg1] = arg2         
print(attendee)         
with open('/home/pi/faceR_prj/temperature.dat','wb') as p:
    pickle.dump(attendee,p) 
Adduser.py
import pickle
import os
import face_recognition
face_encoding = {}
Name = input("Name: ")
Phoneno = input("Phone number :")
NewName = Name
NewName = NewName.replace(" ","_")
os.system('raspistill -w 640 -h 480 -o /home/pi/faceR_prj/oldfiles/'+NewName+'.jpg')
try:
 with open('/home/pi/faceR_prj/encodings.dat','rb') as F:
    face_encoding = pickle.load(F)
except EOFError:
    face_encoding = {}   
try:
 with open('/home/pi/faceR_prj/phonenos.dat','rb') as P:
    Phonenos = pickle.load(P)   
except EOFError:
    Phonenos = {}  
try:
    face_encoding[Name] = face_recognition.face_encodings(face_recognition.load_image_file('/home/pi/faceR_prj/oldfiles/'+NewName+'.jpg'))[0]
    Phonenos[Name] = Phoneno
except IndexError:
    print("NO FACE DETECTED...")
    quit()
with open('/home/pi/faceR_prj/encodings.dat','wb') as f:
    pickle.dump(face_encoding,f)
with open('/home/pi/faceR_prj/phonenos.dat','wb') as p:
    pickle.dump(Phonenos,p)   
os.system('sudo rm -r /home/pi/faceR_prj/oldfiles/'+NewName+'.jpg')
print("User added to database")
Arduino Code
#include <Wire.h>
#include <LiquidCrystal_I2C.h>
#include <Adafruit_MLX90614.h>
#include <Servo.h>
int trigpin1=12;
int echopin1=11;
int trigpin2=7;
int echopin2=6;
int trigpin3=8;
int echopin3=9;
int servopin = 3;
int pingtraveltime;
int pingvalue;
int Length;
Adafruit_MLX90614 mlx = Adafruit_MLX90614();
LiquidCrystal_I2C lcd(0x3F,20,4);
Servo servo ;
int ultrasonic(int trigpin,int echopin)
{
   digitalWrite(trigpin,LOW);
  delayMicroseconds(10);
  digitalWrite(trigpin,HIGH);
  delayMicroseconds(10);
  digitalWrite(trigpin,LOW);
  delayMicroseconds(10);
  pingtraveltime=pulseIn(echopin,HIGH);
  delay(25);
  delay(1000);
  return pingtraveltime;
}
void setup() {
  // put your setup code here, to run once:
 pinMode(trigpin1,OUTPUT);
 pinMode(echopin1,INPUT);
 pinMode(trigpin2,OUTPUT);
 pinMode(echopin2,INPUT);
 pinMode(trigpin3,OUTPUT);
 pinMode(echopin3,INPUT);
 Serial.begin(9600);
 mlx.begin(); 
 servo.attach(servopin);
 servo.write(0);
 lcd.init();
 lcd.backlight();
 lcd.init();
 lcd.setCursor(1,1);
 lcd.print("Booting");
}
void loop() {
  String incoming_data;
 while(Serial.available() == 0 )
     {}     
      incoming_data =  Serial.readStringUntil("\n");
  Length = incoming_data.length()   ;
 if(incoming_data[Length-1] == ',')
 {
  while(1)
  {
    pingvalue = ultrasonic(trigpin1,echopin1) ;
   if(pingvalue > 100 && pingvalue < 5000)
    {
     Serial.println('1');
     break;
    }
   else
    {
     Serial.println('s');
    }
  }
 }
  while(Serial.available() == 0 )
     {}     
      incoming_data =  Serial.readStringUntil("\n");
  Length = incoming_data.length();
  int count = 0;
  if(incoming_data[Length-1]=='*')
    {
   while(count<7)
   { 
       pingvalue = ultrasonic(trigpin2,echopin2);
        if(pingvalue > 100 && pingvalue < 2000)
         {
           double temperature_in_c = mlx.readObjectTempC();
           delay(500);
           String value = String(temperature_in_c);
           Serial.println(value);
           delay(2000);
           break; 
         }
         else
          delay(1000);
         count++;
    }
  count =0;
  while(Serial.available() == 0 )
     {}     
      incoming_data =  Serial.readStringUntil("\n");
  Length = incoming_data.length();
  if(incoming_data[Length-1]=='#')
  {
   while(count<7)
   { 
       pingvalue = ultrasonic(trigpin3,echopin3);
        if(pingvalue > 100 && pingvalue < 2000)
         {
           servo.write(0);
           delay(1000);
           servo.write(140);
           delay(1000);
           servo.write(0);
           delay(1000);
           Serial.println("STOP");
           break; 
         }
        else
          delay(1000);
         count++;
    } 
  }
 }
}